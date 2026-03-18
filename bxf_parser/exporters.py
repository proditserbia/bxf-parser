"""
Export normalized events to CSV and XLSX.
"""
from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import List

from .models import COLUMNS, NormalizedEvent

logger = logging.getLogger(__name__)


def export_csv(rows: List[NormalizedEvent], out_path: Path) -> None:
    """Write rows to a UTF-8 CSV file with BOM for Excel compatibility."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            d = row.as_dict()
            # Convert booleans to readable strings
            d["is_graphics"] = "TRUE" if d["is_graphics"] else "FALSE"
            d["is_live"] = "TRUE" if d["is_live"] else "FALSE"
            d["is_main"] = "TRUE" if d["is_main"] else "FALSE"
            writer.writerow(d)
    logger.info("CSV written: %s (%d rows)", out_path, len(rows))


def export_xlsx(rows: List[NormalizedEvent], out_path: Path) -> None:
    """Write rows to an XLSX workbook using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error(
            "openpyxl is not installed — cannot write XLSX. "
            "Run: pip install openpyxl"
        )
        return

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Alternate row fill colors for readability
    fill_even = PatternFill(fill_type="solid", fgColor="EAF0FB")

    class_colors = {
        "PROGRAMME": "D5E8D4",
        "PROGRAMME_CONTAINER": "82B366",
        "GRAPHICS": "FFE6CC",
        "LIVE_INPUT": "DAE8FC",
        "PLAYOUT": "D5E8D4",
        "OTHER": "F8CECC",
    }

    for row_idx, row in enumerate(rows, start=2):
        d = row.as_dict()
        event_class = d.get("event_class", "OTHER")
        row_color = class_colors.get(event_class)

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = d[col_name]
            if isinstance(value, bool):
                value = "TRUE" if value else "FALSE"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_color:
                cell.fill = PatternFill(fill_type="solid", fgColor=row_color)
            elif row_idx % 2 == 0:
                cell.fill = fill_even

    # Auto-size columns (capped)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(col_name),
            *(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(2, min(len(rows) + 2, 202))
            ),
            0,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(str(out_path))
    logger.info("XLSX written: %s (%d rows)", out_path, len(rows))
